#!/usr/bin/env python3
"""Convert bid evaluation JSON into a formatted 5-sheet Excel workbook.

Reads a structured evaluation JSON (produced by Claude Code during bid-evaluator
Step 3) and generates:
  1. Bid Comparison — normalized pricing matrix with adjustment breakdown
  2. Price Summary — base bid, adjustments, adjusted total, exposure ranges
  3. Exclusion Detail — per-bidder exclusions with risk level scoring
  4. Qualification Summary — per-bidder qualifications and responsibility
  5. Recommendation — ranked bidders with award recommendation

Usage:
  python export_bid_evaluation.py input.json output.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

from shared import safe_output_path


def export_bid_evaluation(input_path, output="Bid_Evaluation.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    with open(input_path) as f:
        data = json.load(f)

    bidders = data.get("bidders", [])
    scope = data.get("scope_description", "")
    project = data.get("project_name", "")
    baseline = data.get("scope_baseline", {})
    recommendation = data.get("recommendation", {})

    if not bidders:
        print("ERROR: No bidders found in input JSON")
        sys.exit(1)

    wb = Workbook()
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    hdr_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid",
    )
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill = PatternFill(
        start_color="F2F7FC", end_color="F2F7FC", fill_type="solid",
    )
    totals_fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid",
    )
    critical_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid",
    )
    significant_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid",
    )
    minor_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid",
    )
    money_fmt = '#,##0.00'
    wrap_top = Alignment(wrap_text=True, vertical="top")

    def write_header(ws, row, headers, widths):
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = w

    bidder_names = [b.get("company_name", f"Bidder {i+1}")
                    for i, b in enumerate(bidders)]

    # -------------------------------------------------------------------
    # Sheet 1: Bid Comparison (normalized pricing matrix)
    # -------------------------------------------------------------------
    ws = wb.active
    ws.title = "Bid Comparison"

    ws["A1"] = f"BID EVALUATION — {scope}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Project: {project}" if project else ""
    ws["A2"].font = Font(italic=True, color="666666")

    headers = ["Item"] + bidder_names
    widths = [35] + [20] * len(bidders)
    write_header(ws, 4, headers, widths)

    # Row data
    comp_rows = [
        ("Submitted Base Bid", "raw_bid"),
        ("Adjustment Total", "_adj_total"),
        ("Adjusted Base Bid", "_adjusted_base"),
    ]
    row = 5
    for label, key in comp_rows:
        ws.cell(row=row, column=1, value=label).border = thin
        ws.cell(row=row, column=1).font = Font(bold=True)
        for c, b in enumerate(bidders, 2):
            cell = ws.cell(row=row, column=c)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if key == "raw_bid":
                cell.value = b.get("raw_bid", "")
            elif key == "_adj_total":
                adjs = b.get("adjustments", [])
                cell.value = sum(a.get("amount", 0) for a in adjs)
            elif key == "_adjusted_base":
                raw = b.get("raw_bid", 0) or 0
                adj = sum(a.get("amount", 0) for a in b.get("adjustments", []))
                cell.value = raw + adj
            if isinstance(cell.value, (int, float)):
                cell.number_format = money_fmt
        if label == "Adjusted Base Bid":
            for c in range(1, 2 + len(bidders)):
                ws.cell(row=row, column=c).fill = totals_fill
        row += 1

    # Adjustment breakdown
    row += 1
    ws.cell(row=row, column=1, value="ADJUSTMENT BREAKDOWN").font = Font(
        bold=True, size=11,
    )
    row += 1

    # Collect all unique adjustment descriptions
    all_adj_descs = []
    for b in bidders:
        for adj in b.get("adjustments", []):
            desc = adj.get("description", "")
            if desc and desc not in all_adj_descs:
                all_adj_descs.append(desc)

    for desc in all_adj_descs:
        ws.cell(row=row, column=1, value=desc).border = thin
        ws.cell(row=row, column=1).alignment = wrap_top
        for c, b in enumerate(bidders, 2):
            cell = ws.cell(row=row, column=c)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            found = None
            for adj in b.get("adjustments", []):
                if adj.get("description", "") == desc:
                    found = adj.get("amount", 0)
                    break
            cell.value = found if found is not None else "N/A"
            if isinstance(cell.value, (int, float)):
                cell.number_format = money_fmt
        row += 1

    # Scope coverage summary
    row += 1
    ws.cell(row=row, column=1, value="SCOPE COVERAGE").font = Font(
        bold=True, size=11,
    )
    row += 1
    coverage_stats = ["INCLUDED", "EXCLUDED", "SILENT", "PARTIAL", "DIFFERENT"]
    for status in coverage_stats:
        ws.cell(row=row, column=1, value=f"{status} items").border = thin
        for c, b in enumerate(bidders, 2):
            cell = ws.cell(row=row, column=c)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            cmap = b.get("coverage_map", {})
            count = sum(
                1 for v in cmap.values() if v.upper() == status
            )
            cell.value = count
            if status == "SILENT" and count > 0:
                cell.fill = critical_fill
            elif status == "EXCLUDED" and count > 0:
                cell.fill = significant_fill
        row += 1

    # -------------------------------------------------------------------
    # Sheet 2: Price Summary
    # -------------------------------------------------------------------
    ws_price = wb.create_sheet(title="Price Summary")
    ws_price["A1"] = "PRICE SUMMARY"
    ws_price["A1"].font = Font(bold=True, size=14)

    headers = [
        "Bidder", "Base Bid", "Adjustments", "Adjusted Total",
        "Rank", "% vs Low", "Exposure Range",
    ]
    widths = [25, 18, 18, 18, 8, 12, 20]
    write_header(ws_price, 3, headers, widths)

    # Calculate adjusted totals and rank
    adjusted_totals = []
    for b in bidders:
        raw = b.get("raw_bid", 0) or 0
        adj = sum(a.get("amount", 0) for a in b.get("adjustments", []))
        adjusted_totals.append(raw + adj)

    low_bid = min(adjusted_totals) if adjusted_totals else 1
    ranked = sorted(range(len(bidders)), key=lambda i: adjusted_totals[i])
    rank_map = {idx: rank + 1 for rank, idx in enumerate(ranked)}

    row = 4
    for i, b in enumerate(bidders):
        raw = b.get("raw_bid", 0) or 0
        adj = sum(a.get("amount", 0) for a in b.get("adjustments", []))
        total = raw + adj
        pct = ((total - low_bid) / low_bid * 100) if low_bid else 0
        exposure = b.get("exposure_range", "")

        ws_price.cell(row=row, column=1, value=bidder_names[i]).border = thin
        for col, val in [(2, raw), (3, adj), (4, total)]:
            cell = ws_price.cell(row=row, column=col, value=val)
            cell.border = thin
            cell.number_format = money_fmt
        ws_price.cell(row=row, column=5, value=rank_map[i]).border = thin
        ws_price.cell(row=row, column=5).alignment = Alignment(
            horizontal="center",
        )
        cell_pct = ws_price.cell(
            row=row, column=6,
            value=f"+{pct:.1f}%" if pct > 0 else "Low",
        )
        cell_pct.border = thin
        cell_pct.alignment = Alignment(horizontal="center")
        ws_price.cell(row=row, column=7, value=exposure).border = thin

        if rank_map[i] == 1:
            for c in range(1, 8):
                ws_price.cell(row=row, column=c).fill = minor_fill
        if i % 2 == 1:
            for c in range(1, 8):
                cell = ws_price.cell(row=row, column=c)
                if not cell.fill or cell.fill.start_color.rgb == "00000000":
                    cell.fill = alt_fill
        row += 1

    # -------------------------------------------------------------------
    # Sheet 3: Exclusion Detail
    # -------------------------------------------------------------------
    ws_exc = wb.create_sheet(title="Exclusion Detail")
    ws_exc["A1"] = "EXCLUSION RISK ANALYSIS"
    ws_exc["A1"].font = Font(bold=True, size=14)
    ws_exc["A2"] = (
        "Risk levels: CRITICAL (spec-required, >1% or >$5K) | "
        "SIGNIFICANT ($2K-$5K) | MINOR (<$2K) | INFO (no cost impact)"
    )
    ws_exc["A2"].font = Font(italic=True, color="666666")

    headers = ["Bidder", "Excluded Item", "Risk Level", "Est. Value", "Notes"]
    widths = [22, 45, 14, 16, 35]
    write_header(ws_exc, 4, headers, widths)

    risk_fills = {
        "CRITICAL": critical_fill,
        "SIGNIFICANT": significant_fill,
        "MINOR": minor_fill,
    }

    row = 5
    for i, b in enumerate(bidders):
        for exc in b.get("exclusion_scores", []):
            ws_exc.cell(row=row, column=1, value=bidder_names[i]).border = thin
            ws_exc.cell(
                row=row, column=2, value=exc.get("item", ""),
            ).border = thin
            ws_exc.cell(row=row, column=2).alignment = wrap_top

            level = exc.get("level", "INFO").upper()
            cell_level = ws_exc.cell(row=row, column=3, value=level)
            cell_level.border = thin
            cell_level.alignment = Alignment(horizontal="center")
            if level in risk_fills:
                cell_level.fill = risk_fills[level]

            cell_val = ws_exc.cell(
                row=row, column=4, value=exc.get("estimated_value", ""),
            )
            cell_val.border = thin
            if isinstance(exc.get("estimated_value"), (int, float)):
                cell_val.number_format = money_fmt

            ws_exc.cell(
                row=row, column=5, value=exc.get("notes", ""),
            ).border = thin
            ws_exc.cell(row=row, column=5).alignment = wrap_top
            row += 1

    # -------------------------------------------------------------------
    # Sheet 4: Qualification Summary
    # -------------------------------------------------------------------
    ws_qual = wb.create_sheet(title="Qualification Summary")
    ws_qual["A1"] = "QUALIFICATION & RESPONSIBILITY ASSESSMENT"
    ws_qual["A1"].font = Font(bold=True, size=14)

    headers = [
        "Factor", *bidder_names,
    ]
    widths = [25] + [22] * len(bidders)
    write_header(ws_qual, 3, headers, widths)

    qual_factors = [
        "bonding", "insurance", "licensing", "experience",
        "capacity", "dbe_mbe", "bid_completeness", "addenda_acknowledged",
        "safety_record", "past_performance",
    ]
    factor_labels = {
        "bonding": "Bonding",
        "insurance": "Insurance",
        "licensing": "Licensing",
        "experience": "Experience",
        "capacity": "Capacity",
        "dbe_mbe": "DBE/MBE",
        "bid_completeness": "Bid Completeness",
        "addenda_acknowledged": "Addenda Acknowledged",
        "safety_record": "Safety Record",
        "past_performance": "Past Performance",
    }

    row = 4
    for factor in qual_factors:
        label = factor_labels.get(factor, factor.replace("_", " ").title())
        ws_qual.cell(row=row, column=1, value=label).border = thin
        ws_qual.cell(row=row, column=1).font = Font(bold=True)
        for c, b in enumerate(bidders, 2):
            quals = b.get("qualifications", {})
            val = quals.get(factor, "N/A")
            cell = ws_qual.cell(row=row, column=c, value=str(val))
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if row % 2 == 0:
            for c in range(1, 2 + len(bidders)):
                ws_qual.cell(row=row, column=c).fill = alt_fill
        row += 1

    # Qualifications / conditions text
    row += 1
    ws_qual.cell(
        row=row, column=1, value="BID CONDITIONS / QUALIFICATIONS",
    ).font = Font(bold=True, size=11)
    row += 1

    for i, b in enumerate(bidders):
        conditions = b.get("conditions", [])
        if conditions:
            ws_qual.cell(
                row=row, column=1, value=bidder_names[i],
            ).font = Font(bold=True)
            row += 1
            for cond in conditions:
                ws_qual.cell(row=row, column=1, value="  \u2022").border = thin
                ws_qual.cell(row=row, column=2, value=cond).border = thin
                ws_qual.cell(row=row, column=2).alignment = wrap_top
                row += 1
            row += 1

    # -------------------------------------------------------------------
    # Sheet 5: Recommendation
    # -------------------------------------------------------------------
    ws_rec = wb.create_sheet(title="Recommendation")
    ws_rec["A1"] = "AWARD RECOMMENDATION"
    ws_rec["A1"].font = Font(bold=True, size=14)
    ws_rec["A2"] = (
        "For PE/PM review — this is an analytical recommendation, "
        "not an award decision."
    )
    ws_rec["A2"].font = Font(italic=True, color="666666")

    row = 4
    rec_fields = [
        ("Recommended Bidder", recommendation.get("recommended_bidder", "")),
        ("Adjusted Bid Amount", recommendation.get("adjusted_amount", "")),
        ("Rank", recommendation.get("rank", "")),
        ("Rationale", recommendation.get("rationale", "")),
    ]
    for label, value in rec_fields:
        ws_rec.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_rec.cell(row=row, column=1).border = thin
        cell = ws_rec.cell(row=row, column=2, value=value)
        cell.border = thin
        cell.alignment = wrap_top
        if isinstance(value, (int, float)):
            cell.number_format = money_fmt
        row += 1

    # Conditions for award
    conditions = recommendation.get("conditions", [])
    if conditions:
        row += 1
        ws_rec.cell(
            row=row, column=1, value="CONDITIONS FOR AWARD",
        ).font = Font(bold=True, size=11)
        row += 1
        for cond in conditions:
            ws_rec.cell(row=row, column=1, value="\u2022").border = thin
            ws_rec.cell(row=row, column=2, value=cond).border = thin
            ws_rec.cell(row=row, column=2).alignment = wrap_top
            row += 1

    # Items needing PE attention
    pe_items = recommendation.get("pe_attention_items", [])
    if pe_items:
        row += 1
        ws_rec.cell(
            row=row, column=1, value="ITEMS REQUIRING PE/PM ATTENTION",
        ).font = Font(bold=True, size=11)
        row += 1
        for item in pe_items:
            ws_rec.cell(row=row, column=1, value="\u2022").border = thin
            ws_rec.cell(row=row, column=2, value=item).border = thin
            ws_rec.cell(row=row, column=2).alignment = wrap_top
            row += 1

    # Ranked bidder summary
    row += 1
    ws_rec.cell(
        row=row, column=1, value="BIDDER RANKING",
    ).font = Font(bold=True, size=11)
    row += 1
    headers = ["Rank", "Bidder", "Adjusted Total", "Key Risk"]
    widths = [8, 25, 18, 40]
    write_header(ws_rec, row, headers, widths)
    row += 1

    ranked_bidders = recommendation.get("ranked_bidders", [])
    for rb in ranked_bidders:
        ws_rec.cell(row=row, column=1, value=rb.get("rank", "")).border = thin
        ws_rec.cell(row=row, column=1).alignment = Alignment(
            horizontal="center",
        )
        ws_rec.cell(
            row=row, column=2, value=rb.get("company", ""),
        ).border = thin
        cell_t = ws_rec.cell(
            row=row, column=3, value=rb.get("adjusted_total", ""),
        )
        cell_t.border = thin
        if isinstance(rb.get("adjusted_total"), (int, float)):
            cell_t.number_format = money_fmt
        ws_rec.cell(
            row=row, column=4, value=rb.get("key_risk", ""),
        ).border = thin
        ws_rec.cell(row=row, column=4).alignment = wrap_top
        row += 1

    # Set column widths on recommendation sheet
    ws_rec.column_dimensions["A"].width = 25
    ws_rec.column_dimensions["B"].width = 50

    # Save
    out = safe_output_path(output)
    wb.save(str(out))
    print(f"OK: {out} ({len(bidders)} bidders evaluated)")
    return str(out)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert bid evaluation JSON to formatted Excel workbook",
    )
    parser.add_argument("input", help="Path to evaluation JSON file")
    parser.add_argument(
        "output", nargs="?", default="Bid_Evaluation.xlsx",
        help="Output Excel file path",
    )
    args = parser.parse_args()
    export_bid_evaluation(args.input, args.output)
