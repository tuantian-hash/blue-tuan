#!/usr/bin/env python3
"""Submittal Log Excel Export — RIGID script.

Claude Code calls this with structured JSON data. This script does not
make any judgment calls — it formats the data identically every time.

Input:  JSON file with submittal_items[], qa_sections[], and project_info{}
Output: Formatted .xlsx workbook with three sheets:
  - Submittal Log: all items with confidence color-coding
  - Summary: pivot-style counts by section, type, and confidence
  - Extraction QA: per-section extraction quality metrics

Usage:
    construction-python export_submittal_log.py --data input.json --output log.xlsx
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from shared import safe_output_path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Color Palette ──────────────────────────────────────────────────

COLORS = {
    "header_bg":     "2C3E50",  # dark slate
    "header_fg":     "FFFFFF",
    "row_even":      "F8F9FA",
    "row_odd":       "FFFFFF",
    "high":          "D4EDDA",  # green
    "medium":        "FFF3CD",  # yellow
    "low":           "FFE0B2",  # orange
    "flagged":       "F8D7DA",  # red
    "high_font":     "155724",
    "medium_font":   "856404",
    "low_font":      "E65100",
    "flagged_font":  "721C24",
    "summary_header": "34495E",
    "summary_accent": "3498DB",
    "border":        "DEE2E6",
}

THIN_BORDER = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

HEADER_FONT = Font(name="Arial", bold=True, color=COLORS["header_fg"], size=10)
HEADER_FILL = PatternFill("solid", fgColor=COLORS["header_bg"])
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")

CONFIDENCE_STYLES = {
    "HIGH":    (COLORS["high"],    COLORS["high_font"]),
    "MEDIUM":  (COLORS["medium"],  COLORS["medium_font"]),
    "LOW":     (COLORS["low"],     COLORS["low_font"]),
    "FLAGGED": (COLORS["flagged"], COLORS["flagged_font"]),
}


# ── Schema ─────────────────────────────────────────────────────────

LOG_COLUMNS = [
    ("Spec Section",          14),
    ("Spec Title",            28),
    ("Submittal No.",         16),
    ("Submittal Type",        18),
    ("Submittal Description", 60),
    ("Article Reference",     14),
    ("Action/Informational",  14),
    ("Confidence",            12),
    ("Flag Reason",           36),
    ("Extraction Method",     14),
    ("Notes",                 36),
]

QA_COLUMNS = [
    ("Spec Section",          14),
    ("Spec Title",            28),
    ("Extraction Method",     16),
    ("Quality Rating",        14),
    ("Failure Modes",         36),
    ("Repair Attempted",      14),
    ("Items Extracted",       14),
    ("Flagged Items",         14),
]


# ── Helpers ────────────────────────────────────────────────────────

def apply_header_row(ws, columns):
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def style_body_cell(cell, row_idx, center=False):
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN if center else WRAP_ALIGN
    bg = COLORS["row_even"] if row_idx % 2 == 0 else COLORS["row_odd"]
    cell.fill = PatternFill("solid", fgColor=bg)


def style_confidence_cell(cell, value):
    if value in CONFIDENCE_STYLES:
        bg, fg = CONFIDENCE_STYLES[value]
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="Arial", size=10, bold=True, color=fg)
        cell.alignment = CENTER_ALIGN


# ── Sheet 1: Submittal Log ────────────────────────────────────────

def write_submittal_log(wb, items):
    ws = wb.active
    ws.title = "Submittal Log"
    ws.sheet_properties.tabColor = "2C3E50"
    apply_header_row(ws, LOG_COLUMNS)

    field_order = [
        "spec_section", "spec_title", "submittal_no", "submittal_type",
        "description", "article_ref", "action_informational",
        "confidence", "flag_reason", "extraction_method", "notes"
    ]
    center_cols = {1, 3, 6, 7, 8, 10}  # 1-indexed columns to center

    for row_idx, item in enumerate(items, 2):
        for col_idx, field in enumerate(field_order, 1):
            value = item.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_body_cell(cell, row_idx, center=(col_idx in center_cols))

            if field == "confidence":
                style_confidence_cell(cell, str(value).upper())

    ws.print_title_rows = "1:1"
    return ws


# ── Sheet 2: Summary ──────────────────────────────────────────────

def write_summary(wb, items, project_info):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "3498DB"

    title_font = Font(name="Arial", bold=True, size=14, color=COLORS["summary_header"])
    subtitle_font = Font(name="Arial", bold=True, size=11, color=COLORS["summary_accent"])
    label_font = Font(name="Arial", size=10, bold=True)
    value_font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 14

    row = 1
    ws.cell(row=row, column=2, value="Submittal Log Summary").font = title_font
    row += 1
    proj_name = project_info.get("project_name", "")
    if proj_name:
        ws.cell(row=row, column=2, value=proj_name).font = Font(
            name="Arial", size=11, color="666666"
        )
        row += 1
    ws.cell(
        row=row, column=2,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ).font = Font(name="Arial", size=9, color="999999")
    row += 2

    # ── Overview stats
    total = len(items)
    action = sum(1 for i in items if i.get("action_informational", "").lower() == "action")
    info = sum(1 for i in items if i.get("action_informational", "").lower() == "informational")
    flagged = sum(
        1 for i in items if str(i.get("confidence", "")).upper() in ("LOW", "FLAGGED")
    )
    sections = len(set(i.get("spec_section", "") for i in items))

    ws.cell(row=row, column=2, value="Overview").font = subtitle_font
    row += 1
    for label, val in [
        ("Total Submittal Items", total),
        ("Action Submittals", action),
        ("Informational Submittals", info),
        ("Items Needing Review (Low/Flagged)", flagged),
        ("Spec Sections with Submittals", sections),
    ]:
        ws.cell(row=row, column=2, value=label).font = label_font
        ws.cell(row=row, column=3, value=val).font = value_font
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        row += 1

    row += 1

    # ── By Confidence
    ws.cell(row=row, column=2, value="By Confidence Level").font = subtitle_font
    row += 1
    for level in ["HIGH", "MEDIUM", "LOW", "FLAGGED"]:
        count = sum(1 for i in items if str(i.get("confidence", "")).upper() == level)
        ws.cell(row=row, column=2, value=level).font = label_font
        c = ws.cell(row=row, column=3, value=count)
        c.font = value_font
        c.alignment = CENTER_ALIGN
        if level in CONFIDENCE_STYLES:
            bg, fg = CONFIDENCE_STYLES[level]
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=bg)
            c.fill = PatternFill("solid", fgColor=bg)
        row += 1

    row += 1

    # ── By Submittal Type (two-column layout with By Section)
    ws.cell(row=row, column=2, value="By Submittal Type").font = subtitle_font
    ws.cell(row=row, column=5, value="By Spec Section").font = subtitle_font
    type_start = row + 1
    section_start = row + 1

    type_counts = {}
    for i in items:
        t = i.get("submittal_type", "Unknown")
        type_counts[t] = type_counts.get(t, 0) + 1
    for t in sorted(type_counts, key=type_counts.get, reverse=True):
        ws.cell(row=type_start, column=2, value=t).font = value_font
        c = ws.cell(row=type_start, column=3, value=type_counts[t])
        c.font = value_font
        c.alignment = CENTER_ALIGN
        type_start += 1

    section_counts = {}
    for i in items:
        key = f"{i.get('spec_section', '')} - {i.get('spec_title', '')}"
        section_counts[key] = section_counts.get(key, 0) + 1
    for s in sorted(section_counts):
        ws.cell(row=section_start, column=5, value=s).font = value_font
        c = ws.cell(row=section_start, column=6, value=section_counts[s])
        c.font = value_font
        c.alignment = CENTER_ALIGN
        section_start += 1

    ws.print_title_rows = "1:1"
    return ws


# ── Sheet 3: Extraction QA ────────────────────────────────────────

def write_qa_sheet(wb, qa_sections):
    ws = wb.create_sheet("Extraction QA")
    ws.sheet_properties.tabColor = "E67E22"
    apply_header_row(ws, QA_COLUMNS)

    field_order = [
        "spec_section", "spec_title", "extraction_method",
        "quality_rating", "failure_modes", "repair_attempted",
        "items_extracted", "flagged_items"
    ]
    center_cols = {1, 3, 4, 6, 7, 8}

    for row_idx, section in enumerate(qa_sections, 2):
        for col_idx, field in enumerate(field_order, 1):
            value = section.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_body_cell(cell, row_idx, center=(col_idx in center_cols))

            if field == "quality_rating":
                rating = str(value).upper()
                if rating == "GOOD":
                    cell.fill = PatternFill("solid", fgColor=COLORS["high"])
                    cell.font = Font(name="Arial", size=10, bold=True, color=COLORS["high_font"])
                elif rating == "DEGRADED":
                    cell.fill = PatternFill("solid", fgColor=COLORS["medium"])
                    cell.font = Font(name="Arial", size=10, bold=True, color=COLORS["medium_font"])
                elif rating == "POOR":
                    cell.fill = PatternFill("solid", fgColor=COLORS["flagged"])
                    cell.font = Font(name="Arial", size=10, bold=True, color=COLORS["flagged_font"])

    ws.print_title_rows = "1:1"
    return ws


# ── Main ───────────────────────────────────────────────────────────

def export(input_path, output_path):
    with open(input_path, "r") as f:
        data = json.load(f)

    items = data.get("submittal_items", [])
    qa_sections = data.get("qa_sections", [])
    project_info = data.get("project_info", {})

    if not items:
        print("WARNING: No submittal items found in input data.")

    wb = Workbook()
    write_submittal_log(wb, items)
    write_summary(wb, items, project_info)
    write_qa_sheet(wb, qa_sections)

    out = safe_output_path(output_path)
    wb.save(str(out))
    print(f"OK: {out}")
    print(f"  Submittal items: {len(items)}")
    print(f"  QA sections:     {len(qa_sections)}")
    print(f"  Confidence: "
          f"HIGH={sum(1 for i in items if str(i.get('confidence', '')).upper() == 'HIGH')} "
          f"MED={sum(1 for i in items if str(i.get('confidence', '')).upper() == 'MEDIUM')} "
          f"LOW={sum(1 for i in items if str(i.get('confidence', '')).upper() == 'LOW')} "
          f"FLAG={sum(1 for i in items if str(i.get('confidence', '')).upper() == 'FLAGGED')}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export submittal log JSON to formatted Excel")
    parser.add_argument("--data", required=True, help="Input JSON file")
    parser.add_argument("--output", "-o", default="Submittal_Log.xlsx", help="Output Excel file")
    args = parser.parse_args()

    if not Path(args.data).exists():
        print(f"ERROR: Input file not found: {args.data}")
        sys.exit(1)

    export(args.data, args.output)
