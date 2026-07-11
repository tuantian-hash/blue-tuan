#!/usr/bin/env python3
"""Convert extracted bid data into a formatted comparison Excel workbook.

Reads per-bidder JSON files from a directory and produces a multi-tab workbook:
  - Comparison Summary (cross-sheet formulas referencing per-bidder tabs)
  - Per-bidder detail tabs (line items, SUM subtotal, reconciliation check)
  - Exclusions & Qualifications
  - Scope Gaps (items appearing in some bids but not others)
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from shared import safe_output_path

_VE_RE = re.compile(
    r'(DEDUCT|ADD|CREDIT)\s+\$?([\d,]+(?:\.\d{2})?)',
    re.IGNORECASE,
)


def _parse_ve_deducts(qualifications):
    """Extract VE/deduct/add amounts from qualification strings.

    Returns list of (description, signed_amount) tuples.
    Deducts are negative, adds are positive. Informational only —
    these are NOT subtracted from the base bid total.
    """
    results = []
    for q in qualifications:
        m = _VE_RE.search(q)
        if m:
            kind = m.group(1).upper()
            raw = m.group(2).replace(",", "")
            amount = float(raw)
            if kind == "DEDUCT":
                amount = -amount
            results.append((q, amount))
    return results


def _quote_tab(name):
    """Quote sheet name for cross-sheet Excel formulas."""
    escaped = name.replace("'", "''")
    return f"'{escaped}'"


def bid_comparison_to_xlsx(
    data_dir, output="Bid_Comparison.xlsx", scope="", project="",
    scope_matrix=None,
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    # Load all bid JSON files
    data_path = Path(data_dir)
    bids = []
    for f in sorted(data_path.glob("*.json")):
        with open(f) as fh:
            bids.append(json.load(fh))

    if not bids:
        print(f"ERROR: No bid JSON files found in {data_dir}")
        sys.exit(1)

    wb = Workbook()
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    thick_top = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("medium"), bottom=Side("thin"),
    )
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    totals_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    money_fmt = '#,##0.00'
    wrap_top = Alignment(wrap_text=True, vertical="top")

    def write_header(ws, row, headers, widths):
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = w

    # Extended amount is column 7 (G) in the 8-column line item layout
    ext_col = 7
    num_li_cols = 8

    def _totals_row(ws, row, label, ext_value, note="", border=thin):
        """Write a styled totals row across the line item columns (1-8).

        The monetary value goes in the Extended column (7/G).
        """
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=1).font = Font(bold=True)
        for c in range(2, ext_col):
            ws.cell(row=row, column=c, value="").border = border
        cell = ws.cell(row=row, column=ext_col, value=ext_value)
        cell.border = border
        cell.font = Font(bold=True)
        if isinstance(ext_value, str) and ext_value.startswith("="):
            cell.number_format = money_fmt
        elif isinstance(ext_value, (int, float)):
            cell.number_format = money_fmt
        ws.cell(row=row, column=num_li_cols, value=note).border = border
        for c in range(1, num_li_cols + 1):
            ws.cell(row=row, column=c).fill = totals_fill

    bidder_names = [
        b.get("company_name", f"Bidder {i+1}") for i, b in enumerate(bids)
    ]

    # -------------------------------------------------------------------
    # Per-Bidder Detail Tabs (built first for cross-sheet references)
    # -------------------------------------------------------------------
    bidder_refs = []  # {tab_title, subtotal_row, submitted_row, recon_row}

    for i, bid in enumerate(bids):
        name = bidder_names[i]
        tab_name = name[:28] + "..." if len(name) > 31 else name
        ws_bid = wb.create_sheet(title=tab_name)

        ws_bid["A1"] = name
        ws_bid["A1"].font = Font(bold=True, size=14)
        ws_bid["A2"] = bid.get("contact_name", "")
        ws_bid["A3"] = (
            f"{bid.get('contact_phone', '')}  |  {bid.get('contact_email', '')}"
        )
        ws_bid["A4"] = f"Bid Date: {bid.get('bid_date', 'N/A')}"
        ws_bid["A4"].font = Font(bold=True)

        # Line items
        row = 6
        ws_bid.cell(
            row=row, column=1, value="LINE ITEMS (as-submitted)",
        ).font = Font(bold=True, size=11)
        row += 1
        li_headers = [
            "#", "Spec Section", "Description", "Qty", "Unit",
            "Unit Price", "Extended", "Notes",
        ]
        li_widths = [6, 14, 42, 10, 8, 14, 16, 28]
        write_header(ws_bid, row, li_headers, li_widths)
        row += 1

        first_item_row = row
        has_numeric = False
        for j, item in enumerate(bid.get("line_items", []), 1):
            if isinstance(item, str):
                spec, desc, qty, unit = "", item, "", ""
                uprice, ext, note = "", "", ""
            else:
                spec = item.get("spec_section", "")
                desc = item.get("description", "")
                qty = item.get("qty")
                unit = item.get("unit", "")
                uprice = item.get("unit_price")
                ext = item.get(
                    "extended_price", item.get("amount", ""),
                )
                note = item.get("notes", "")

            ws_bid.cell(row=row, column=1, value=j).border = thin
            ws_bid.cell(row=row, column=2, value=spec or "").border = thin
            cell_d = ws_bid.cell(row=row, column=3, value=desc)
            cell_d.border = thin
            cell_d.alignment = wrap_top
            cell_q = ws_bid.cell(
                row=row, column=4,
                value=qty if qty is not None else "",
            )
            cell_q.border = thin
            if isinstance(qty, (int, float)):
                cell_q.number_format = '#,##0'
            ws_bid.cell(
                row=row, column=5, value=unit or "",
            ).border = thin
            cell_up = ws_bid.cell(
                row=row, column=6,
                value=uprice if uprice is not None else "",
            )
            cell_up.border = thin
            if isinstance(uprice, (int, float)):
                cell_up.number_format = money_fmt
            cell_ext = ws_bid.cell(
                row=row, column=ext_col,
                value=ext if ext is not None else "",
            )
            cell_ext.border = thin
            if isinstance(ext, (int, float)):
                cell_ext.number_format = money_fmt
                has_numeric = True
            cell_n = ws_bid.cell(
                row=row, column=num_li_cols, value=note or "",
            )
            cell_n.border = thin
            cell_n.alignment = wrap_top
            if j % 2 == 0:
                for c in range(1, num_li_cols + 1):
                    ws_bid.cell(row=row, column=c).fill = alt_fill
            row += 1
        last_item_row = row - 1

        # --- Totals block ---
        row += 1  # blank separator

        subtotal_row = row
        sub_val = (
            f"=SUM(G{first_item_row}:G{last_item_row})"
            if has_numeric else "N/A"
        )
        _totals_row(ws_bid, row, "Line Item Subtotal", sub_val, border=thick_top)
        row += 1

        submitted_row = row
        _totals_row(
            ws_bid, row, "Submitted Bid Total",
            bid.get("base_bid_amount", ""),
            note="(from PDF)",
        )
        ws_bid.cell(row=row, column=num_li_cols).font = Font(
            italic=True, color="666666",
        )
        row += 1

        recon_row = row
        recon_val = (
            f"=G{subtotal_row}-G{submitted_row}" if has_numeric else "N/A"
        )
        recon_note = (
            f'=IF(ABS(G{recon_row})<1,"MATCH","DISCREPANCY")'
            if has_numeric else ""
        )
        _totals_row(ws_bid, row, "Reconciliation", recon_val, note=recon_note)
        row += 1

        # Alternates
        if bid.get("alternates"):
            row += 1
            ws_bid.cell(
                row=row, column=1, value="ALTERNATES",
            ).font = Font(bold=True, size=11)
            row += 1
            for alt in bid["alternates"]:
                ws_bid.cell(row=row, column=1, value=alt.get("name", "")).border = thin
                cell_ad = ws_bid.cell(row=row, column=2, value=alt.get("description", ""))
                cell_ad.border = thin
                cell_ad.alignment = wrap_top
                cell_aa = ws_bid.cell(row=row, column=3, value=alt.get("amount", ""))
                cell_aa.border = thin
                if isinstance(alt.get("amount"), (int, float)):
                    cell_aa.number_format = money_fmt
                row += 1

        # VE / Deducts (informational — NOT subtracted from base bid)
        ve_items = _parse_ve_deducts(bid.get("qualifications", []))
        if ve_items:
            row += 1
            ws_bid.cell(
                row=row, column=1,
                value="VALUE ENGINEERING / DEDUCTS (optional, not applied to total)",
            ).font = Font(bold=True, size=11)
            row += 1
            ve_headers = ["#", "Description", "Amount", ""]
            ve_widths = [6, 50, 18, 30]
            write_header(ws_bid, row, ve_headers, ve_widths)
            row += 1
            first_ve = row
            for k, (desc, amt) in enumerate(ve_items, 1):
                ws_bid.cell(row=row, column=1, value=k).border = thin
                cell_vd = ws_bid.cell(row=row, column=2, value=desc)
                cell_vd.border = thin
                cell_vd.alignment = wrap_top
                cell_va = ws_bid.cell(row=row, column=3, value=amt)
                cell_va.border = thin
                cell_va.number_format = money_fmt
                ws_bid.cell(row=row, column=4, value="").border = thin
                row += 1
            last_ve = row - 1

            _totals_row(
                ws_bid, row, "VE Subtotal (if accepted)",
                f"=SUM(C{first_ve}:C{last_ve})",
                border=thick_top,
            )
            row += 1

        # Inclusions
        if bid.get("scope_inclusions"):
            row += 1
            ws_bid.cell(
                row=row, column=1, value="SCOPE INCLUSIONS",
            ).font = Font(bold=True, size=11)
            row += 1
            for inc in bid["scope_inclusions"]:
                ws_bid.cell(row=row, column=1, value="•").border = thin
                cell_inc = ws_bid.cell(row=row, column=2, value=inc)
                cell_inc.border = thin
                cell_inc.alignment = wrap_top
                row += 1

        # Exclusions
        if bid.get("scope_exclusions"):
            row += 1
            ws_bid.cell(
                row=row, column=1, value="SCOPE EXCLUSIONS",
            ).font = Font(bold=True, size=11)
            row += 1
            for exc in bid["scope_exclusions"]:
                ws_bid.cell(row=row, column=1, value="•").border = thin
                cell_exc = ws_bid.cell(row=row, column=2, value=exc)
                cell_exc.border = thin
                cell_exc.alignment = wrap_top
                row += 1

        # Qualifications (excluding VE items already shown above)
        ve_texts = {d for d, _ in ve_items} if ve_items else set()
        remaining_quals = [
            q for q in bid.get("qualifications", []) if q not in ve_texts
        ]
        if remaining_quals:
            row += 1
            ws_bid.cell(
                row=row, column=1, value="QUALIFICATIONS / CONDITIONS",
            ).font = Font(bold=True, size=11)
            row += 1
            for q in remaining_quals:
                ws_bid.cell(row=row, column=1, value="•").border = thin
                cell_q = ws_bid.cell(row=row, column=2, value=q)
                cell_q.border = thin
                cell_q.alignment = wrap_top
                row += 1

        bidder_refs.append({
            "tab_title": tab_name,
            "subtotal_row": subtotal_row,
            "submitted_row": submitted_row,
            "recon_row": recon_row,
        })

    # -------------------------------------------------------------------
    # Comparison Summary (cross-sheet formulas)
    # -------------------------------------------------------------------
    ws = wb.active
    ws.title = "Comparison Summary"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(bids))
    ws["A1"] = f"BID COMPARISON — {scope}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Project: {project}" if project else ""
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d')}  |  {len(bids)} bids"
    )
    ws["A3"].font = Font(italic=True, color="666666")

    headers = ["Item", ""] + bidder_names
    widths = [30, 2] + [20] * len(bids)
    write_header(ws, 5, headers, widths)

    # Formula-linked summary rows
    formula_rows = [
        ("Submitted Bid Total", "submitted_row"),
        ("Line Item Subtotal", "subtotal_row"),
        ("Reconciliation", "recon_row"),
    ]

    row = 6
    for label, ref_key in formula_rows:
        ws.cell(row=row, column=1, value=label).border = thin
        ws.cell(row=row, column=1).font = Font(bold=True)
        for c, ref in enumerate(bidder_refs, 3):
            cell = ws.cell(row=row, column=c)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            qs = _quote_tab(ref["tab_title"])
            cell.value = f"={qs}!G{ref[ref_key]}"
            cell.number_format = money_fmt
        row += 1

    row += 1  # separator

    # Static info rows
    info_rows = [
        ("Bond Included", [b.get("bond_included", "") for b in bids]),
        ("Schedule Duration", [b.get("schedule_duration", "") for b in bids]),
        ("Payment Terms", [b.get("payment_terms", "") for b in bids]),
        ("Bid Date", [b.get("bid_date", "") for b in bids]),
        ("Bid Validity", [b.get("bid_validity_period", "") for b in bids]),
    ]

    # Alternates in summary
    all_alt_names = []
    for b in bids:
        for alt in b.get("alternates", []):
            aname = alt.get("name", alt.get("description", ""))
            if aname and aname not in all_alt_names:
                all_alt_names.append(aname)
    for alt_name in all_alt_names:
        vals = []
        for b in bids:
            found = None
            for alt in b.get("alternates", []):
                if alt.get("name", alt.get("description", "")) == alt_name:
                    found = alt.get("amount", "N/A")
                    break
            vals.append(found if found is not None else "Not included")
        info_rows.append((f"Alt: {alt_name}", vals))

    for label, values in info_rows:
        ws.cell(row=row, column=1, value=label).border = thin
        ws.cell(row=row, column=1).font = Font(bold=True)
        for c, v in enumerate(values, 3):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if isinstance(v, (int, float)):
                cell.number_format = money_fmt
        row += 1

    # -------------------------------------------------------------------
    # Exclusions & Qualifications
    # -------------------------------------------------------------------
    ws_exc = wb.create_sheet(title="Exclusions & Qualifications")
    ws_exc["A1"] = "EXCLUSIONS & QUALIFICATIONS — ALL BIDDERS"
    ws_exc["A1"].font = Font(bold=True, size=14)

    headers = ["Bidder", "Type", "Item"]
    widths = [25, 16, 60]
    write_header(ws_exc, 3, headers, widths)
    row = 4
    for idx, bid in enumerate(bids):
        bname = bidder_names[idx]
        for exc in bid.get("scope_exclusions", []):
            ws_exc.cell(row=row, column=1, value=bname).border = thin
            ws_exc.cell(row=row, column=2, value="Exclusion").border = thin
            cell_ei = ws_exc.cell(row=row, column=3, value=exc)
            cell_ei.border = thin
            cell_ei.alignment = wrap_top
            row += 1
        for q in bid.get("qualifications", []):
            ws_exc.cell(row=row, column=1, value=bname).border = thin
            ws_exc.cell(row=row, column=2, value="Qualification").border = thin
            cell_qi = ws_exc.cell(row=row, column=3, value=q)
            cell_qi.border = thin
            cell_qi.alignment = wrap_top
            row += 1

    # -------------------------------------------------------------------
    # Scope Gaps — matrix (items × bidders)
    # -------------------------------------------------------------------
    ws_gaps = wb.create_sheet(title="Scope Gaps")
    ws_gaps["A1"] = "SCOPE INCLUSION / EXCLUSION MATRIX"
    ws_gaps["A1"].font = Font(bold=True, size=14)
    ws_gaps["A2"] = (
        "Items from all bids' inclusion and exclusion lists. "
        "Engineer to verify if gaps are real or terminology differences."
    )
    ws_gaps["A2"].font = Font(italic=True, color="666666")

    inc_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    exc_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Load scope matrix (Claude Code normalized) or fall back to text-match
    matrix_data = None
    if scope_matrix:
        sm_path = Path(scope_matrix)
        if sm_path.exists():
            with open(sm_path) as f:
                matrix_data = json.load(f)

    # Short bidder labels for column headers
    short_names = []
    for bn in bidder_names:
        parts = bn.replace(",", "").split()
        short_names.append(parts[0] if parts else bn[:10])

    headers = ["Scope Item"] + short_names
    widths = [50] + [14] * len(bids)
    write_header(ws_gaps, 4, headers, widths)

    row = 5
    if matrix_data and matrix_data.get("groups"):
        # Normalized matrix from Claude Code
        for group in matrix_data["groups"]:
            cell_desc = ws_gaps.cell(row=row, column=1, value=group["label"])
            cell_desc.border = thin
            cell_desc.alignment = wrap_top
            for c, bname in enumerate(bidder_names, 2):
                cell = ws_gaps.cell(row=row, column=c)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")
                status = group.get("bidders", {}).get(bname, "not_mentioned")
                if status == "included":
                    cell.value = "✓"
                    cell.fill = inc_fill
                elif status == "excluded":
                    cell.value = "✗"
                    cell.fill = exc_fill
                    cell.font = Font(color="CC0000")
                else:
                    cell.value = ""
            row += 1
    else:
        # Fallback: text-match matrix from raw bid data
        scope_items = {}
        for idx, bid in enumerate(bids):
            bname = bidder_names[idx]
            for inc in bid.get("scope_inclusions", []):
                key = inc.strip().lower()
                if key not in scope_items:
                    scope_items[key] = {
                        "text": inc, "included": set(), "excluded": set(),
                    }
                scope_items[key]["included"].add(bname)
            for exc in bid.get("scope_exclusions", []):
                key = exc.strip().lower()
                if key not in scope_items:
                    scope_items[key] = {
                        "text": exc, "included": set(), "excluded": set(),
                    }
                scope_items[key]["excluded"].add(bname)

        sorted_items = sorted(
            scope_items.values(),
            key=lambda x: len(x["included"]) + len(x["excluded"]),
            reverse=True,
        )
        for item in sorted_items:
            cell_desc = ws_gaps.cell(row=row, column=1, value=item["text"])
            cell_desc.border = thin
            cell_desc.alignment = wrap_top
            for c, bname in enumerate(bidder_names, 2):
                cell = ws_gaps.cell(row=row, column=c)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")
                if bname in item["included"] and bname in item["excluded"]:
                    cell.value = "✓/✗"
                    cell.font = Font(color="CC0000", bold=True)
                elif bname in item["included"]:
                    cell.value = "✓"
                    cell.fill = inc_fill
                elif bname in item["excluded"]:
                    cell.value = "✗"
                    cell.fill = exc_fill
                    cell.font = Font(color="CC0000")
                else:
                    cell.value = ""
            row += 1

    if row == 5:
        ws_gaps.cell(
            row=5, column=1,
            value="No scope items detected from inclusion/exclusion lists.",
        )
        ws_gaps.cell(row=5, column=1).font = Font(italic=True, color="666666")

    # Save
    out = safe_output_path(output)
    wb.save(str(out))
    print(f"OK: {out} ({len(bids)} bids)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert bid data to comparison Excel workbook",
    )
    parser.add_argument(
        "--data", required=True,
        help="Directory containing per-bidder JSON files",
    )
    parser.add_argument("--output", "-o", default="Bid_Comparison.xlsx")
    parser.add_argument("--scope", default="", help="Scope of work description")
    parser.add_argument("--project", default="", help="Project name")
    parser.add_argument("--scope-matrix", default=None, help="Path to scope_matrix.json")
    args = parser.parse_args()
    bid_comparison_to_xlsx(
        args.data, args.output, args.scope, args.project,
        scope_matrix=args.scope_matrix,
    )
