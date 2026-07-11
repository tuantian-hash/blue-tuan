#!/usr/bin/env python3
"""QTO Excel Export — RIGID script.

Claude Code calls this with structured JSON data matching the QTO output
format from references/qto-output-format.md. This script does not make
any judgment calls — it formats the data identically every time.

Input:  JSON file with project{}, scope{}, totals{}, line_items[], etc.
Output: Formatted .xlsx workbook with four sheets:
  - QTO Summary: one row per line item with counts and confidence
  - Instance Detail: every detection record with full provenance
  - Type Definitions: type defs applied and their room targets
  - Completeness: coverage metrics, gap sheets, and issues

Usage:
    construction-python qto_to_xlsx.py \\
        --data qto_data.json \\
        --project "Holabird Academy" \\
        --scope "Room Tags" \\
        --output "QTO_Room_Tags.xlsx"
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from shared import safe_output_path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ── Color Palette ──────────────────────────────────────────────────

COLORS = {
    "header_bg":     "2C3E50",
    "header_fg":     "FFFFFF",
    "row_even":      "F8F9FA",
    "row_odd":       "FFFFFF",
    "high":          "D4EDDA",
    "medium":        "FFF3CD",
    "low":           "FFE0B2",
    "flagged":       "F8D7DA",
    "high_font":     "155724",
    "medium_font":   "856404",
    "low_font":      "E65100",
    "flagged_font":  "721C24",
    "title_bg":      "1A3C5E",
    "subtitle_bg":   "34495E",
    "accent":        "3498DB",
    "border":        "DEE2E6",
    "dedup_bg":      "E9ECEF",
}

THIN_BORDER = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

HEADER_FONT = Font(name="Arial", bold=True, color=COLORS["header_fg"], size=10)
HEADER_FILL = PatternFill("solid", fgColor=COLORS["header_bg"])
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="top")

TITLE_FONT = Font(name="Arial", bold=True, color=COLORS["header_fg"], size=14)
TITLE_FILL = PatternFill("solid", fgColor=COLORS["title_bg"])
SUBTITLE_FONT = Font(name="Arial", bold=True, color=COLORS["header_fg"], size=11)
SUBTITLE_FILL = PatternFill("solid", fgColor=COLORS["subtitle_bg"])

CONFIDENCE_STYLES = {
    "high":    (COLORS["high"],    COLORS["high_font"]),
    "medium":  (COLORS["medium"],  COLORS["medium_font"]),
    "low":     (COLORS["low"],     COLORS["low_font"]),
    "flagged": (COLORS["flagged"], COLORS["flagged_font"]),
}


# ── Column Definitions ────────────────────────────────────────────

SUMMARY_COLUMNS = [
    ("Element",        24),
    ("Designation",    14),
    ("Sheet Instances", 14),
    ("Derived",        12),
    ("Deduplicated",   12),
    ("Building Qty",   14),
    ("Avg Confidence", 14),
    ("Notes",          40),
]

DETAIL_COLUMNS = [
    ("Sheet",          12),
    ("Tag Text",       24),
    ("Designation",    14),
    ("Type",           16),
    ("View Type",      14),
    ("Room",           10),
    ("Provenance",     16),
    ("Confidence",     12),
    ("Status",         12),
    ("Detection ID",   36),
    ("Element IDs",    40),
]

TYPEDEF_COLUMNS = [
    ("Source Detail",      16),
    ("Parent View Label",  24),
    ("Elements Defined",   40),
    ("Applied To Rooms",   40),
    ("Applied Count",      14),
    ("Derivation Method",  18),
]

COMPLETENESS_COLUMNS = [
    ("Metric",     24),
    ("Value",      40),
]


# ── Helpers ────────────────────────────────────────────────────────

def apply_header(ws, columns, row=1):
    """Write header row and set column widths."""
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(columns))}{row}"


def style_body_cell(cell, row_idx, align=None):
    """Apply standard body styling with alternating row shading."""
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = align or WRAP_ALIGN
    fill_color = COLORS["row_even"] if row_idx % 2 == 0 else COLORS["row_odd"]
    cell.fill = PatternFill("solid", fgColor=fill_color)


def style_confidence_cell(cell, confidence_val):
    """Color-code a cell based on confidence level."""
    if isinstance(confidence_val, (int, float)):
        if confidence_val >= 0.8:
            level = "high"
        elif confidence_val >= 0.5:
            level = "medium"
        elif confidence_val > 0:
            level = "low"
        else:
            level = "flagged"
    elif isinstance(confidence_val, str):
        level = confidence_val.lower()
    else:
        return

    if level in CONFIDENCE_STYLES:
        bg, fg = CONFIDENCE_STYLES[level]
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="Arial", size=10, color=fg, bold=True)


def write_title_row(ws, text, col_count, row=1):
    """Merge and style a title row spanning all columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30
    return row + 1


# ── Sheet Builders ─────────────────────────────────────────────────

def build_summary_sheet(ws, data, project_name, scope_name):
    """Build the QTO Summary sheet."""
    ws.title = "QTO Summary"

    # Title row
    row = write_title_row(
        ws,
        f"QTO Summary — {project_name} — {scope_name}",
        len(SUMMARY_COLUMNS),
    )

    # Totals sub-header
    totals = data.get("totals", {})
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(SUMMARY_COLUMNS))
    cell = ws.cell(
        row=row, column=1,
        value=(
            f"Sheet Instances: {totals.get('sheet_instances', 0)}  |  "
            f"Derived: {totals.get('derived_instances', 0)}  |  "
            f"Dedup: {totals.get('deduplicated', 0)}  |  "
            f"Building Qty: {totals.get('building_quantity', 0)}"
        ),
    )
    cell.font = SUBTITLE_FONT
    cell.fill = SUBTITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    # Headers
    apply_header(ws, SUMMARY_COLUMNS, row=row)
    row += 1

    # Line items
    for item in data.get("line_items", []):
        element = item.get("element", "")
        designation = item.get("designation", "")
        sheet_inst = item.get("sheet_instances", 0)
        derived = item.get("derived_instances", 0)
        dedup = item.get("deduplicated", 0)
        bldg_qty = item.get("building_qty", 0)

        # Compute average confidence from instance details
        details = item.get("instance_details", [])
        avg_conf = 0.0
        if details:
            confs = [d.get("confidence", 0) for d in details if isinstance(d.get("confidence"), (int, float))]
            avg_conf = sum(confs) / len(confs) if confs else 0.0

        notes = ""
        if dedup > 0:
            notes = f"{dedup} dedup'd (direct takes precedence)"

        values = [element, designation, sheet_inst, derived, dedup, bldg_qty, avg_conf, notes]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            style_body_cell(cell, row)

            # Number formatting for quantity columns
            if col_idx in (3, 4, 5, 6):
                cell.number_format = "#,##0"
                cell.alignment = RIGHT_ALIGN
            elif col_idx == 7:
                cell.number_format = "0.00"
                cell.alignment = CENTER_ALIGN
                style_confidence_cell(cell, val)

        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col_idx, key in [(3, "sheet_instances"), (4, "derived_instances"), (5, "deduplicated"), (6, "building_quantity")]:
        cell = ws.cell(row=row, column=col_idx, value=totals.get(key, 0))
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = THIN_BORDER
        cell.number_format = "#,##0"
        cell.alignment = RIGHT_ALIGN


def build_detail_sheet(ws, data):
    """Build the Instance Detail sheet."""
    ws.title = "Instance Detail"
    apply_header(ws, DETAIL_COLUMNS)
    row = 2

    for item in data.get("line_items", []):
        for det in item.get("instance_details", []):
            values = [
                det.get("sheet", ""),
                det.get("tag_text", item.get("element", "")),
                item.get("designation", ""),
                det.get("type", "element_instance"),
                det.get("view_type", "floor_plan"),
                det.get("room", ""),
                det.get("provenance", det.get("type", "")),
                det.get("confidence", 0),
                det.get("status", "pending"),
                det.get("detection_id", ""),
                ", ".join(det.get("element_ids", [])) if det.get("element_ids") else "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                style_body_cell(cell, row)
                if col_idx == 8:
                    cell.number_format = "0.00"
                    cell.alignment = CENTER_ALIGN
                    style_confidence_cell(cell, val)
                elif col_idx == 9:
                    cell.alignment = CENTER_ALIGN

            row += 1

        # Derived instances
        for der in item.get("derived_details", []):
            values = [
                der.get("target_sheet", ""),
                item.get("element", ""),
                item.get("designation", ""),
                "derived_instance",
                "derived",
                der.get("target_room", ""),
                f"from {der.get('source_detail', '')} ({der.get('derivation_method', '')})",
                "",
                der.get("status", "pending"),
                der.get("derived_id", ""),
                "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                style_body_cell(cell, row)
                if col_idx == 9:
                    cell.alignment = CENTER_ALIGN

            # Mark deduplicated rows with muted background
            if der.get("deduplicated"):
                for col_idx in range(1, len(DETAIL_COLUMNS) + 1):
                    ws.cell(row=row, column=col_idx).fill = PatternFill("solid", fgColor=COLORS["dedup_bg"])

            row += 1


def build_typedef_sheet(ws, data):
    """Build the Type Definitions sheet."""
    ws.title = "Type Definitions"
    apply_header(ws, TYPEDEF_COLUMNS)
    row = 2

    for td in data.get("type_definitions_applied", []):
        elements_str = "; ".join(
            f"{e.get('tag_text', '')} x{e.get('count_per_instance', 1)}"
            for e in td.get("elements_defined", [])
        )
        rooms_str = ", ".join(str(r) for r in td.get("applied_to_rooms", []))

        values = [
            td.get("source_detail", ""),
            td.get("parent_view_label", ""),
            elements_str,
            rooms_str,
            td.get("applied_count", 0),
            td.get("derivation_method", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            style_body_cell(cell, row)
            if col_idx == 5:
                cell.number_format = "#,##0"
                cell.alignment = CENTER_ALIGN

        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No type definitions applied").font = Font(name="Arial", italic=True, size=10)


def build_completeness_sheet(ws, data, project_name, scope_name):
    """Build the Completeness sheet."""
    ws.title = "Completeness"

    row = write_title_row(
        ws,
        f"Completeness Report — {project_name} — {scope_name}",
        len(COMPLETENESS_COLUMNS),
    )
    apply_header(ws, COMPLETENESS_COLUMNS, row=row)
    row += 1

    scope = data.get("scope", {})
    totals = data.get("totals", {})
    comp = data.get("completeness", {})

    metrics = [
        ("Tag Type", scope.get("tag_type", "")),
        ("CSI Division", scope.get("csi_division", "")),
        ("Sheets Scanned", str(len(scope.get("sheets_scanned", [])))),
        ("Sheets with Detections", str(len(scope.get("sheets_with_detections", [])))),
        ("Sheets with Zero", ", ".join(scope.get("sheets_with_zero", [])) or "None"),
        ("", ""),
        ("Building Quantity", str(totals.get("building_quantity", 0))),
        ("Sheet Instances", str(totals.get("sheet_instances", 0))),
        ("Derived Instances", str(totals.get("derived_instances", 0))),
        ("Deduplicated", str(totals.get("deduplicated", 0))),
        ("", ""),
        ("Schedule Reference", comp.get("schedule_reference", "N/A")),
        ("Expected Count", str(comp.get("expected_count", "N/A"))),
        ("Detected Count", str(comp.get("detected_count", "N/A"))),
        ("Coverage", f"{comp.get('coverage_pct', 0):.1f}%" if comp.get("coverage_pct") else "N/A"),
        ("Gap Sheets", ", ".join(comp.get("gap_sheets", [])) or "None"),
        ("Gap Notes", comp.get("gap_notes", "")),
    ]

    for metric, value in metrics:
        if not metric:
            row += 1
            continue
        m_cell = ws.cell(row=row, column=1, value=metric)
        v_cell = ws.cell(row=row, column=2, value=value)
        style_body_cell(m_cell, row)
        style_body_cell(v_cell, row)
        m_cell.font = Font(name="Arial", bold=True, size=10)
        row += 1

    # Issues section
    issues = data.get("issues", [])
    if issues:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value="Issues & Warnings")
        cell.font = SUBTITLE_FONT
        cell.fill = SUBTITLE_FILL
        row += 1

        for issue in issues:
            severity = issue.get("severity", "info").upper()
            msg = issue.get("message", "")
            action = issue.get("suggested_action", "")
            text = f"[{severity}] {msg}"
            if action:
                text += f" — Action: {action}"

            cell = ws.cell(row=row, column=1, value=text)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            style_body_cell(cell, row)

            if severity == "WARNING":
                cell.fill = PatternFill("solid", fgColor=COLORS["medium"])
                cell.font = Font(name="Arial", size=10, color=COLORS["medium_font"])
            elif severity == "ERROR":
                cell.fill = PatternFill("solid", fgColor=COLORS["flagged"])
                cell.font = Font(name="Arial", size=10, color=COLORS["flagged_font"])

            row += 1

    # Generated timestamp
    row += 1
    ts = data.get("project", {}).get("generated_at", datetime.now().isoformat())
    ws.cell(row=row, column=1, value=f"Generated: {ts}").font = Font(
        name="Arial", italic=True, size=9, color="808080",
    )


# ── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="QTO Excel Export")
    parser.add_argument("--data", required=True, help="Path to QTO JSON file")
    parser.add_argument("--project", default="", help="Project name for headers")
    parser.add_argument("--scope", default="", help="Scope label (e.g., 'Room Tags')")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    # Load data
    data_path = Path(args.data)
    if not data_path.exists():
        print(f"ERROR: Data file not found: {data_path}", file=sys.stderr)
        sys.exit(1)

    with open(data_path, "r") as f:
        data = json.load(f)

    # Validate expected schema keys — the skill must produce exact key names
    required_keys = ["line_items", "totals", "scope"]
    missing = [k for k in required_keys if k not in data]
    if missing:
        print(f"WARNING: Missing expected keys: {missing}. Output may be empty.", file=sys.stderr)
        print(f"  Found keys: {list(data.keys())}", file=sys.stderr)
        print(f"  See references/qto-output-format.md for required schema.", file=sys.stderr)

    project_name = args.project or data.get("project", {}).get("name", "Unknown Project")
    scope_name = args.scope or data.get("scope", {}).get("tag_type", "Tags")

    # Build workbook
    wb = Workbook()

    build_summary_sheet(wb.active, data, project_name, scope_name)
    build_detail_sheet(wb.create_sheet(), data)
    build_typedef_sheet(wb.create_sheet(), data)
    build_completeness_sheet(wb.create_sheet(), data, project_name, scope_name)

    # Save
    out = safe_output_path(args.output)
    wb.save(str(out))
    print(f"OK: {out}")


if __name__ == "__main__":
    main()
