#!/usr/bin/env python3
"""Parse an anchored schedule Excel file and compute a reconciliation changeset.

Reads the _agentcm_meta sheet for anchor IDs, reads the data sheet with _row_key
column, queries the current DB state via psql, and outputs a JSON changeset
suitable for POST to /api/projects/:id/schedules/reconcile.

Usage:
    python xlsx_to_changeset.py --excel edited.xlsx --query-command "PGPASSWORD=reader psql ..." [--output changeset.json]

If --output is omitted, prints JSON to stdout.
"""

import argparse
import json
import subprocess
import sys
from pathlib import Path


def read_meta(wb):
    """Extract reconciliation anchors from the _agentcm_meta sheet."""
    if "_agentcm_meta" not in wb.sheetnames:
        print("ERROR: No _agentcm_meta sheet found. This Excel was not exported by AgentCM.", file=sys.stderr)
        sys.exit(1)

    meta_ws = wb["_agentcm_meta"]
    meta = {}
    for row in meta_ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] and row[1]:
            meta[str(row[0]).strip()] = str(row[1]).strip()

    required = ["schedule_id", "schedule_type"]
    for key in required:
        if key not in meta or not meta[key]:
            print(f"ERROR: _agentcm_meta missing required field: {key}", file=sys.stderr)
            sys.exit(1)

    return meta


def read_data(wb, meta):
    """Read the data sheet, returning headers and rows keyed by _row_key."""
    # Data sheet is the first visible sheet
    data_ws = None
    for name in wb.sheetnames:
        if name != "_agentcm_meta":
            data_ws = wb[name]
            break

    if data_ws is None:
        print("ERROR: No data worksheet found.", file=sys.stderr)
        sys.exit(1)

    # Find header row (row with _row_key in column A)
    header_row = None
    for row_idx, row in enumerate(data_ws.iter_rows(min_row=1, max_col=1, values_only=True), 1):
        if row[0] and str(row[0]).strip() == "_row_key":
            header_row = row_idx
            break

    if header_row is None:
        print("ERROR: Could not find _row_key header in column A.", file=sys.stderr)
        sys.exit(1)

    # Read column headers (B onward)
    headers = []
    for col_idx in range(2, data_ws.max_column + 1):
        val = data_ws.cell(row=header_row, column=col_idx).value
        if val is not None:
            headers.append(str(val).strip())
        else:
            break

    # Read data rows
    excel_rows = {}  # row_key -> {col: value}
    new_rows = []    # rows with empty _row_key (PE-added rows)

    for row_idx in range(header_row + 1, data_ws.max_row + 1):
        row_key = data_ws.cell(row=row_idx, column=1).value

        cells = {}
        has_data = False
        for col_offset, header in enumerate(headers):
            val = data_ws.cell(row=row_idx, column=col_offset + 2).value
            cell_str = str(val).strip() if val is not None else None
            cells[header] = cell_str
            if cell_str:
                has_data = True

        if not has_data:
            continue  # skip fully empty rows

        if row_key and str(row_key).strip():
            excel_rows[str(row_key).strip()] = cells
        else:
            new_rows.append(cells)

    return headers, excel_rows, new_rows


def query_db_state(query_command, schedule_id):
    """Query current schedule state from the database via psql."""
    # Get schedule columns
    sql_columns = (
        f"SELECT columns FROM schedules WHERE id = '{schedule_id}'"
    )

    # Get current rows with cells
    sql_rows = (
        f"SELECT sr.entity_identifier, sc.column_key, sc.cell_value "
        f"FROM schedule_rows sr "
        f"JOIN schedule_cells sc ON sc.schedule_row_id = sr.id "
        f"WHERE sr.schedule_id = '{schedule_id}' AND sr.is_deleted = FALSE "
        f"ORDER BY sr.entity_identifier, sc.column_key"
    )

    db_rows = {}       # entity_identifier -> {col: value}
    db_columns = set()

    # Execute rows query
    try:
        result = subprocess.run(
            query_command.split() + ["-t", "-A", "-F", "\t", "-c", sql_rows],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0:
            print(f"ERROR: psql query failed: {result.stderr.strip()}", file=sys.stderr)
            sys.exit(1)

        for line in result.stdout.strip().split("\n"):
            if not line.strip():
                continue
            parts = line.split("\t")
            if len(parts) >= 3:
                entity_id, col_key, cell_val = parts[0], parts[1], parts[2] if parts[2] else None
                if entity_id not in db_rows:
                    db_rows[entity_id] = {}
                db_rows[entity_id][col_key] = cell_val
                db_columns.add(col_key)

    except subprocess.TimeoutExpired:
        print("ERROR: Database query timed out.", file=sys.stderr)
        sys.exit(1)

    return db_rows, db_columns


def compute_changeset(meta, excel_headers, excel_rows, new_rows, db_rows, db_columns):
    """Compute the diff between Excel state and DB state."""
    changeset = {
        "scheduleId": meta["schedule_id"],
        "sourceFile": meta.get("source_file", ""),
        "cellUpdates": [],
        "rowAdds": [],
        "rowDeletes": [],
        "columnAdds": [],
        "columnHides": [],
    }

    excel_header_set = set(excel_headers)

    # --- Cell updates: compare matching rows ---
    for row_key, excel_cells in excel_rows.items():
        if row_key not in db_rows:
            continue  # new row (handled below)

        db_cells = db_rows[row_key]
        for col, excel_val in excel_cells.items():
            if col not in db_columns:
                continue  # new column (handled below)
            db_val = db_cells.get(col)

            # Normalize for comparison (treat None and empty string as equivalent)
            norm_excel = (excel_val or "").strip()
            norm_db = (db_val or "").strip()

            if norm_excel != norm_db:
                changeset["cellUpdates"].append({
                    "rowKey": row_key,
                    "columnKey": col,
                    "newValue": excel_val,
                })

    # --- New rows: row_key in Excel but not in DB ---
    for row_key, excel_cells in excel_rows.items():
        if row_key not in db_rows:
            changeset["rowAdds"].append({
                "entityIdentifier": row_key,
                "cells": excel_cells,
            })

    # --- PE-added rows (empty _row_key) ---
    # These need natural key inference — include the first non-empty cell as identifier
    for cells in new_rows:
        # Use first non-empty value as entity identifier
        entity_id = ""
        for header in excel_headers:
            if cells.get(header):
                entity_id = cells[header]
                break
        if entity_id:
            changeset["rowAdds"].append({
                "entityIdentifier": entity_id,
                "cells": cells,
            })

    # --- Deleted rows: row_key in DB but not in Excel ---
    for row_key in db_rows:
        if row_key not in excel_rows:
            changeset["rowDeletes"].append({"rowKey": row_key})

    # --- New columns: in Excel headers but not in DB ---
    for header in excel_headers:
        if header not in db_columns:
            changeset["columnAdds"].append({
                "key": header,
                "header": header,
            })

    # --- Hidden columns: in DB but not in Excel headers ---
    for col in db_columns:
        if col not in excel_header_set:
            changeset["columnHides"].append({"key": col})

    return changeset


def main():
    parser = argparse.ArgumentParser(description="Parse anchored Excel and compute reconciliation changeset")
    parser.add_argument("--excel", required=True, help="Path to the edited Excel file")
    parser.add_argument("--query-command", required=True, help="psql query command from database.yaml")
    parser.add_argument("--output", "-o", help="Output JSON file (default: stdout)")
    args = parser.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    # Read Excel
    wb = load_workbook(args.excel, data_only=True)
    meta = read_meta(wb)
    excel_headers, excel_rows, new_rows = read_data(wb, meta)

    # Query DB state
    db_rows, db_columns = query_db_state(args.query_command, meta["schedule_id"])

    # Compute changeset
    changeset = compute_changeset(meta, excel_headers, excel_rows, new_rows, db_rows, db_columns)

    # Summary
    summary = {
        "cellUpdates": len(changeset["cellUpdates"]),
        "rowAdds": len(changeset["rowAdds"]),
        "rowDeletes": len(changeset["rowDeletes"]),
        "columnAdds": len(changeset["columnAdds"]),
        "columnHides": len(changeset["columnHides"]),
    }
    total = sum(summary.values())

    if total == 0:
        print("No changes detected between Excel and database.", file=sys.stderr)

    # Output
    output_json = json.dumps(changeset, indent=2)
    if args.output:
        Path(args.output).write_text(output_json)
        print(f"OK: {args.output} ({total} changes: {summary})", file=sys.stderr)
    else:
        print(output_json)
        print(f"({total} changes: {summary})", file=sys.stderr)


if __name__ == "__main__":
    main()
