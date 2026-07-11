#!/usr/bin/env python3
"""Convert extracted schedule data (JSON) to a formatted Excel workbook.

Produces an anchored Excel file for reconciliation round-trips:
- Hidden _agentcm_meta sheet with schedule_id, schedule_type, sheet_id
- Hidden _row_key column A with entity identifiers (door number, room number, etc.)
- Data columns start at B, frozen at B2
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from shared import safe_output_path

SCHEMA_VERSION = "1"
AGENTCM_VERSION = "1.0.0"


def schedule_to_xlsx(data_path, output="schedule.xlsx", schedule_type="generic",
                     project="", sheet="", schedule_id="", sheet_id="",
                     row_key_column=""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    with open(data_path) as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = schedule_type.replace("_", " ").title()

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Determine columns and rows from data
    if isinstance(data, list) and len(data) > 0:
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            rows = [list(row.values()) for row in data]
        elif isinstance(data[0], list):
            headers = data[0]
            rows = data[1:]
        else:
            headers = ["Value"]
            rows = [[v] for v in data]
    else:
        print("ERROR: Data must be a non-empty list")
        sys.exit(1)

    # Resolve row key column index (for _row_key population)
    row_key_idx = _resolve_row_key_index(headers, row_key_column, schedule_type)

    # --- Column A: hidden _row_key ---
    start_row = 4
    ws.cell(row=start_row, column=1, value="_row_key")
    ws.cell(row=start_row, column=1).fill = header_fill
    ws.cell(row=start_row, column=1).font = header_font
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=start_row, column=1).border = thin_border

    # --- Project info rows (shifted to B) ---
    ws.merge_cells("B1:F1")
    ws["B1"] = f"Project: {project}"
    ws["B1"].font = Font(bold=True, size=14)
    ws.merge_cells("B2:F2")
    ws["B2"] = f"Source: Sheet {sheet} | Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B2"].font = Font(italic=True, size=10, color="666666")

    # --- Data headers (B onward) ---
    for col, header in enumerate(headers, 2):  # start at column B (2)
        cell = ws.cell(row=start_row, column=col, value=str(header))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # --- Data rows ---
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r, row in enumerate(rows, start_row + 1):
        # Column A: _row_key value
        key_val = row[row_key_idx] if row_key_idx < len(row) else ""
        ws.cell(row=r, column=1, value=key_val)
        ws.cell(row=r, column=1).border = thin_border

        # Columns B onward: data
        for c, val in enumerate(row, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if (r - start_row) % 2 == 0:
                cell.fill = alt_fill

    # Hide column A (_row_key)
    ws.column_dimensions["A"].hidden = True

    # Auto-size data columns (B onward)
    for col in range(2, len(headers) + 2):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(start_row, start_row + len(rows) + 1)
        )
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(max_len + 4, 40)

    # Freeze panes at B5 (after hidden _row_key column + header row 4)
    ws.freeze_panes = "B5"

    # --- Hidden _agentcm_meta sheet ---
    meta = wb.create_sheet("_agentcm_meta")
    meta.sheet_state = "hidden"
    meta_fields = [
        ("schedule_id", schedule_id),
        ("schedule_type", schedule_type),
        ("sheet_id", sheet_id),
        ("exported_at", datetime.now().isoformat()),
        ("schema_version", SCHEMA_VERSION),
        ("agentcm_version", AGENTCM_VERSION),
        ("project", project),
        ("source_sheet", sheet),
        ("row_count", str(len(rows))),
        ("column_count", str(len(headers))),
    ]
    for r, (key, val) in enumerate(meta_fields, 1):
        meta.cell(row=r, column=1, value=key).font = Font(bold=True)
        meta.cell(row=r, column=2, value=str(val))

    out = safe_output_path(output)
    wb.save(str(out))
    print(f"OK: {out} ({len(rows)} rows, {len(headers)} columns)")


def _resolve_row_key_index(headers, row_key_column, schedule_type):
    """Find the column index to use as _row_key value.

    Priority: explicit --row-key-column > natural key by schedule type > first column.
    """
    if row_key_column:
        for i, h in enumerate(headers):
            if h.strip().upper() == row_key_column.strip().upper():
                return i

    # Natural key patterns by schedule type (per SCHEDULE_RECONCILIATION_SPEC §4.5)
    natural_key_patterns = {
        "door": ["DOOR NO", "DOOR NUMBER", "DOOR MARK", "MARK", "NO.", "NO"],
        "finish": ["ROOM NO", "ROOM NUMBER", "ROOM", "RM NO", "RM"],
        "fixture": ["FIXTURE TAG", "FIXTURE MARK", "TAG", "MARK"],
        "equipment": ["EQUIPMENT TAG", "EQUIP TAG", "TAG", "MARK"],
        "plumbing_fixture": ["FIXTURE DESIGNATION", "FIXTURE", "MARK"],
        "window": ["WINDOW NO", "WINDOW MARK", "MARK", "NO."],
        "panel": ["PANEL", "PANEL NO", "PANEL NAME"],
    }

    patterns = natural_key_patterns.get(schedule_type, [])
    header_upper = [h.strip().upper() for h in headers]
    for pattern in patterns:
        for i, h in enumerate(header_upper):
            if h == pattern:
                return i

    return 0  # fallback: first column


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert schedule data to Excel")
    parser.add_argument("--data", required=True, help="JSON data file")
    parser.add_argument("--output", "-o", default="schedule.xlsx")
    parser.add_argument("--type", default="generic", help="Schedule type")
    parser.add_argument("--project", default="", help="Project name")
    parser.add_argument("--sheet", default="", help="Source sheet number")
    parser.add_argument("--schedule-id", default="", help="DB schedule UUID (for reconciliation)")
    parser.add_argument("--sheet-id", default="", help="DB sheet UUID (for reconciliation)")
    parser.add_argument("--row-key-column", default="", help="Column name to use as row key")
    args = parser.parse_args()
    schedule_to_xlsx(
        args.data, args.output, args.type, args.project, args.sheet,
        args.schedule_id, args.sheet_id, args.row_key_column,
    )
